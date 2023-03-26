cd#!/usr/bin/env python3

from typing import Optional
from hdwallet import BIP44HDWallet
from hdwallet.cryptocurrencies import EthereumMainnet
from hdwallet.derivations import BIP44Derivation
from hdwallet.utils import generate_mnemonic
import openpyxl

book = openpyxl.Workbook()
sheet = book.active
sheet['A1'] = "Mnemonic"
sheet['B1'] = "Address"
sheet['C1'] = "Private key"


if input("Скрипт может сгенерировать seed фразы, а может взять уже готовые из файла mnemonic.txt . Сгенерировать новые seed фразы? (y/n): ")=="y":
    # Number of wallets
    N = int(input("Сколько надо кошельков ? : "))

    for i in range(N):
        # Generate english mnemonic words
        MNEMONIC: str = generate_mnemonic(language="english", strength=128)
        # Secret passphrase/password for mnemonic
        PASSPHRASE: Optional[str] = None  # "meherett"

        # Initialize Ethereum mainnet BIP44HDWallet
        bip44_hdwallet: BIP44HDWallet = BIP44HDWallet(cryptocurrency=EthereumMainnet)
        # Get Ethereum BIP44HDWallet from mnemonic
        bip44_hdwallet.from_mnemonic(
            mnemonic=MNEMONIC, language="english", passphrase=PASSPHRASE
        )
        # Clean default BIP44 derivation indexes/paths
        bip44_hdwallet.clean_derivation()
        sheet.cell(row=i+2,column=1).value=bip44_hdwallet.mnemonic()


        # Get Ethereum BIP44HDWallet information's from address index

        # Derivation from Ethereum BIP44 derivation path
        bip44_derivation: BIP44Derivation = BIP44Derivation(
            cryptocurrency=EthereumMainnet, account=0, change=False, address=0
        )
        # Drive Ethereum BIP44HDWallet
        bip44_hdwallet.from_path(path=bip44_derivation)
        # Print address_index, path, address and private_key
        sheet.cell(row=i+2,column=2).value=bip44_hdwallet.address()
        sheet.cell(row=i+2, column=3).value=bip44_hdwallet.private_key()
        # Clean derivation indexes/paths
        bip44_hdwallet.clean_derivation()
else :

    with open("mnemonic.txt", "r") as f:
        mnemonic_list = [row.strip() for row in f]
        
    i = 0
    for MNEMONIC in mnemonic_list:
        
        # Generate english mnemonic words
        #MNEMONIC: str = generate_mnemonic(language="english", strength=128)
        # Secret passphrase/password for mnemonic
        PASSPHRASE: Optional[str] = None  # "meherett"

        # Initialize Ethereum mainnet BIP44HDWallet
        bip44_hdwallet: BIP44HDWallet = BIP44HDWallet(cryptocurrency=EthereumMainnet)
        # Get Ethereum BIP44HDWallet from mnemonic
        bip44_hdwallet.from_mnemonic(
            mnemonic=MNEMONIC, language="english", passphrase=PASSPHRASE
        )
        # Clean default BIP44 derivation indexes/paths
        bip44_hdwallet.clean_derivation()
        sheet.cell(row=i+2,column=1).value=bip44_hdwallet.mnemonic()


        # Get Ethereum BIP44HDWallet information's from address index

        # Derivation from Ethereum BIP44 derivation path
        bip44_derivation: BIP44Derivation = BIP44Derivation(
            cryptocurrency=EthereumMainnet, account=0, change=False, address=0
        )
        # Drive Ethereum BIP44HDWallet
        bip44_hdwallet.from_path(path=bip44_derivation)
        # Print address_index, path, address and private_key
        sheet.cell(row=i+2,column=2).value=bip44_hdwallet.address()
        sheet.cell(row=i+2, column=3).value=bip44_hdwallet.private_key()
        # Clean derivation indexes/paths
        bip44_hdwallet.clean_derivation()

        i=i+1

book.save("MM.xlsx")
book.close()